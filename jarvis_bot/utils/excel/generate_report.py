import openpyxl

from data import config


async def do_file_db(c):
    fileName = 'all_info.xlsx'
    wb = openpyxl.load_workbook(filename=fileName)
    sheet = wb['Пользователи']
    c.execute("select user_id, create_date, tg_first_name, tg_last_name, tg_username, is_live, is_block from users")
    infos = c.fetchall()
    i = 2
    sheet.delete_rows(2, 10000)
    for info in infos:
        user_id, create_date, tg_first_name, tg_last_name, tg_username, is_live, is_block = info
        sheet.cell(row=i, column=1).value = user_id
        sheet.cell(row=i, column=2).value = create_date
        sheet.cell(row=i, column=3).value = tg_first_name
        sheet.cell(row=i, column=4).value = tg_last_name
        sheet.cell(row=i, column=5).value = tg_username
        sheet.cell(row=i, column=6).value = "✅" if is_live else "❌"
        sheet.cell(row=i, column=7).value = "✅" if is_block else "❌"
        i += 1
    sheet = wb['Заказы']
    c.execute("select order_id, date_create, is_urgent, courier_id, order_status, complete_photo, "
              "receipt_address, address_comment, receipt_phone, sender_phone, "
              "date_format(delivery_min_hours, '%H:%i'), date_format(delivery_max_hours, '%H:%i'), "
              "delivery_courier_amount, is_fine, how_much_to_go "
              "from orders")
    infos = c.fetchall()
    i = 2
    sheet.delete_rows(2, 10000)
    for info in infos:
        order_id, date_create, is_urgent, courier_id, order_status, complete_photo, receipt_address, address_comment, \
        receipt_phone, sender_phone, delivery_min_hours, delivery_max_hours, delivery_courier_amount, is_fine, \
        how_much_to_go = info
        c.execute("select full_name from couriers where courier_id = %s", (courier_id,))
        try:
            full_name = c.fetchone()[0]
        except:
            full_name = 'курьер удален'
        sheet.cell(row=i, column=1).value = order_id
        sheet.cell(row=i, column=2).value = date_create
        sheet.cell(row=i, column=3).value = "🔥 СРОЧНО" if is_urgent else "💤 не срочный"
        sheet.cell(row=i, column=4).value = full_name
        sheet.cell(row=i, column=5).value = config.ORDER_STATUSES[order_status][1]
        sheet.cell(row=i, column=6).value = complete_photo
        sheet.cell(row=i, column=7).value = receipt_address
        sheet.cell(row=i, column=8).value = address_comment
        sheet.cell(row=i, column=9).value = receipt_phone
        sheet.cell(row=i, column=10).value = sender_phone
        sheet.cell(row=i, column=11).value = f'с {delivery_min_hours} до {delivery_max_hours}' if not is_urgent else f'ДО {delivery_min_hours}'
        sheet.cell(row=i, column=12).value = delivery_courier_amount
        sheet.cell(row=i, column=13).value = is_fine
        sheet.cell(row=i, column=14).value = how_much_to_go
        i += 1
    sheet = wb['Менеджеры']
    c.execute("select manager_id, full_name from managers")
    infos = c.fetchall()
    i = 2
    sheet.delete_rows(2, 10000)
    for info in infos:
        manager_id, full_name = info
        sheet.cell(row=i, column=1).value = manager_id
        sheet.cell(row=i, column=2).value = full_name
        i += 1
    sheet = wb['Курьеры']
    c.execute("select courier_id, full_name, passport_photo_first, passport_photo_second, card_number, balance, "
              "wa_phone, mobile_phone, is_car, gov_number, car_color from couriers")
    infos = c.fetchall()
    i = 2
    sheet.delete_rows(2, 10000)
    for info in infos:
        courier_id, full_name, passport_photo_first, passport_photo_second, card_number, balance, wa_phone, \
        mobile_phone, is_car, gov_number, car_color = info
        sheet.cell(row=i, column=1).value = courier_id
        sheet.cell(row=i, column=2).value = full_name
        sheet.cell(row=i, column=3).value = passport_photo_first
        sheet.cell(row=i, column=4).value = passport_photo_second
        sheet.cell(row=i, column=5).value = card_number
        sheet.cell(row=i, column=6).value = balance
        sheet.cell(row=i, column=7).value = wa_phone
        sheet.cell(row=i, column=8).value = mobile_phone
        sheet.cell(row=i, column=9).value = "авто 🚗" if is_car else "пеший 🏃"
        sheet.cell(row=i, column=10).value = gov_number
        sheet.cell(row=i, column=11).value = car_color if not car_color else config.CAR_COLORS[car_color][1]
        i += 1
    sheet = wb['Выводы средств']
    c.execute("select withdraw_date, courier_id, withdraw_amount from couriers_withdraws")
    infos = c.fetchall()
    i = 2
    sheet.delete_rows(2, 10000)
    for info in infos:
        withdraw_date, courier_id, withdraw_amount = info
        c.execute("select full_name from couriers where courier_id = %s", (courier_id,))
        try:
            full_name = c.fetchone()[0]
        except:
            full_name = 'курьер удален'
        sheet.cell(row=i, column=1).value = withdraw_date
        sheet.cell(row=i, column=2).value = courier_id
        sheet.cell(row=i, column=3).value = full_name
        sheet.cell(row=i, column=4).value = withdraw_amount
        i += 1
    # сохраняем данные
    wb.save(fileName)
    wb.close()
    return open(fileName, 'rb')
