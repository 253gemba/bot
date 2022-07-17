import openpyxl


async def do_file_db(c):
    fileName = 'all_info.xlsx'
    wb = openpyxl.load_workbook(filename=fileName)
    sheet_online = wb['Онлайн']
    sheet_offline = wb['Оффлайн']
    c.execute("select user_id, date_reg, user_first_name, user_username, full_name, phone_number, email, company_name, "
              "company_position, is_offline, accept_mailing "
              "from users "
              "order by (select count(*) from users_answers where user_id = users.user_id and is_true = 1) desc ")
    infos = c.fetchall()
    i = 2
    sheet_online.delete_rows(2, 10000)
    sheet_offline.delete_rows(2, 10000)
    for info in infos:
        user_id, date_reg, user_first_name, user_username, full_name, phone_number, email, company_name, \
        company_position, is_offline, accept_mailing = info
        if is_offline:
            sheet = sheet_offline
        else:
            sheet = sheet_online
        c.execute("select count(*) from users_answers where user_id = %s", (user_id,))
        count_all = c.fetchone()[0]
        c.execute("select count(*) from users_answers where user_id = %s and is_true = 1", (user_id,))
        count_true = c.fetchone()[0]
        sheet.cell(row=i, column=1).value = user_id
        sheet.cell(row=i, column=2).value = date_reg
        sheet.cell(row=i, column=3).value = user_first_name
        sheet.cell(row=i, column=4).value = user_username
        sheet.cell(row=i, column=5).value = full_name
        sheet.cell(row=i, column=6).value = phone_number
        sheet.cell(row=i, column=7).value = email
        sheet.cell(row=i, column=8).value = company_name
        sheet.cell(row=i, column=9).value = company_position
        sheet.cell(row=i, column=10).value = "оффлайн" if is_offline else "онлайн"
        sheet.cell(row=i, column=11).value = "✅" if accept_mailing else "✖"
        sheet.cell(row=i, column=12).value = f'{count_true}'
        sheet.cell(row=i, column=13).value = f'{count_all}'
        i += 1
    sheet_results = wb['Результаты']
    c.execute("select user_id, "
              "(select full_name from users where users.user_id = users_answers.user_id), "
              "question_id, "
              "(select answer_name from possible_answers "
              "where possible_answers.answer_id = users_answers.answer_id), "
              "is_true "
              "from users_answers")
    infos = c.fetchall()
    i = 2
    sheet_results.delete_rows(2, 10000)
    for info in infos:
        user_id, full_name, question_id, answer_name, is_true = info
        sheet_results.cell(row=i, column=1).value = user_id
        sheet_results.cell(row=i, column=2).value = full_name
        sheet_results.cell(row=i, column=3).value = question_id
        sheet_results.cell(row=i, column=4).value = answer_name
        sheet_results.cell(row=i, column=5).value = "✅" if is_true else "✖"
        i += 1
    # сохраняем данные
    wb.save(fileName)
    wb.close()
    return open(fileName, 'rb')
